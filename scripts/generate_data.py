from pathlib import Path
import json
import re
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / 'Mapping for Acadience.xlsx'
ASSET_MANIFEST = ROOT / 'data' / 'generated' / 'asset-manifest.json'
OUT_FILE = ROOT / 'data' / 'generated' / 'recommendations.json'
TODO_FILE = ROOT / 'docs' / 'TODO.md'

GRADES = ['K', '1', '2']
SEASONS = ['BOY', 'MOY', 'EOY']
BANDS = {
    'red': ['red', 'well below'],
    'yellow': ['yellow', 'below'],
    'green': ['green', 'benchmark'],
    'blue': ['blue', 'above'],
}

CONCEPT_ALIAS = {
    'rhyme': 'Rhyme',
    'syllable': 'Syllable',
    'sound isolation': 'Sound Isolation',
    'sound blending': 'Sound Blending',
    'sound segmenting': 'Sound Segmenting',
    'letter sound correspondence': 'Letter-Sound Knowledge',
    'letter-sound correspondence': 'Letter-Sound Knowledge',
    'decoding': 'Word Building',
    'high frequency words': 'High Frequency Words',
    'high frequency words / word building': 'High Frequency Words',
    'fluency': 'Fluency',
    'comprehension': 'Comprehension',
}

GAME_BY_CONCEPT = {
    'Rhyme': ['rhyming-match', 'rhyming-memory'],
    'Syllable': ['syllable-match'],
    'Sound Isolation': ['fly-swat'],
    'Sound Blending': ['sound-blender'],
    'Sound Segmenting': ['fly-swat'],
    'Letter-Sound Knowledge': ['letter-sound-match'],
    'Word Building': ['word-build-race'],
    'High Frequency Words': ['sight-word-snap'],
    'Fluency': ['speed-read-sprint'],
    'Comprehension': ['story-question-quest'],
}


def norm(s):
    if s is None:
        return ''
    return re.sub(r'\s+', ' ', str(s).strip().lower())


def parse_band(raw):
    t = norm(raw)
    for band, tokens in BANDS.items():
        if any(tok in t for tok in tokens):
            return band
    return None


def first_matching_asset(concept_assets, keywords, ext):
    kws = [norm(k) for k in keywords if k]
    for a in concept_assets:
        if a['ext'] != ext:
            continue
        hay = norm(a['filename'])
        if any(k in hay for k in kws):
            return '/' + a['publicPath']
    return None


def split_labels(value):
    text = str(value or '').strip()
    if not text:
        return []
    parts = re.split(r'[/,+]| and ', text, flags=re.I)
    return [part.strip() for part in parts if part and part.strip()]


def first_global_matching_asset(all_assets, keywords, ext):
    kws = [norm(k) for k in keywords if k]
    for a in all_assets:
        if a['ext'] != ext:
            continue
        hay = norm(a['filename'])
        if any(k and k in hay for k in kws):
            return '/' + a['publicPath']
    return None

assets = json.loads(ASSET_MANIFEST.read_text(encoding='utf-8')) if ASSET_MANIFEST.exists() else []
assets_by_concept = {}
for a in assets:
    assets_by_concept.setdefault(a['concept'], []).append(a)

wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
concept_sheet = wb['Copy of Sheet1']
mapping_sheet = wb['Sheet2']

concept_rows = []
for row in concept_sheet.iter_rows(min_row=2, values_only=True):
    concept = row[0]
    if not concept:
        continue
    concept_name = CONCEPT_ALIAS.get(norm(concept), str(concept).strip())
    concept_rows.append({
        'concept': concept_name,
        'explainer': row[1],
        'videoLabel': row[2],
        'instructionLabel': row[3],
        'cardsLabel': row[4],
    })

concept_meta = {}
for row in concept_rows:
    concept = row['concept']
    concept_assets = assets_by_concept.get(concept, [])
    concept_meta[concept] = {
        'concept': concept,
        'explainer': row['explainer'],
        'video': None,
        'instructionPdf': first_matching_asset(concept_assets, ['instruction', 'games', 'activities', 'tool'], '.pdf'),
        'cardsPdf': first_matching_asset(concept_assets, ['cards', 'tool', 'activities'], '.pdf'),
        'games': GAME_BY_CONCEPT.get(concept, []),
    }

# Create metadata for alias-only concepts that are present in assets but absent in Sheet1 rows.
for concept in set(CONCEPT_ALIAS.values()):
    if concept in concept_meta:
        continue
    concept_assets = assets_by_concept.get(concept, [])
    if not concept_assets:
        continue
    concept_meta[concept] = {
        'concept': concept,
        'explainer': None,
        'video': None,
        'instructionPdf': first_matching_asset(concept_assets, ['instruction', 'games', 'activities', 'tool', 'build-a-word'], '.pdf'),
        'cardsPdf': first_matching_asset(concept_assets, ['cards', 'tool', 'activities', 'letter'], '.pdf'),
        'games': GAME_BY_CONCEPT.get(concept, []),
    }

header = [cell.value for cell in mapping_sheet[1]]
col_map = {}
for idx, label in enumerate(header, start=1):
    n = norm(label)
    if not n:
        continue
    if n in {'focus concept', 'focusconcept'}:
        col_map['concept'] = idx
        continue
    m = re.match(r'^(k|1|2)\s+(boy|moy|eoy)$', n)
    if m:
        col_map[(m.group(1).upper(), m.group(2).upper())] = idx

recommendations = []
missing = []

for row_idx in range(2, mapping_sheet.max_row + 1):
    concept_raw = mapping_sheet.cell(row=row_idx, column=col_map['concept']).value
    if not concept_raw:
        continue
    concept = CONCEPT_ALIAS.get(norm(concept_raw), str(concept_raw).strip())

    for grade in GRADES:
        for season in SEASONS:
            col = col_map.get((grade, season))
            if not col:
                continue
            band = parse_band(mapping_sheet.cell(row=row_idx, column=col).value)
            if not band:
                continue
            cm = concept_meta.get(concept)
            if not cm:
                missing.append(f'Missing concept metadata for {concept} (Sheet2 row {row_idx})')
                continue
            recommendations.append({
                'grade': grade,
                'season': season,
                'measure': 'composite',
                'band': band,
                'recommendedSkillArea': concept,
                'games': cm['games'],
                'resources': {
                    'video': cm['video'],
                    'instructionPdf': cm['instructionPdf'],
                    'cardsPdf': cm['cardsPdf'],
                },
            })

# Validate coverage
for grade in GRADES:
    for season in SEASONS:
        for band in BANDS.keys():
            if not any(r for r in recommendations if r['grade'] == grade and r['season'] == season and r['band'] == band):
                missing.append(f'No mapping row for {grade} {season} {band}')

# Validate assets
for r in recommendations:
    if not r['resources']['instructionPdf']:
        missing.append(f"Missing instruction PDF for {r['recommendedSkillArea']} ({r['grade']} {r['season']} {r['band']})")

OUT_FILE.parent.mkdir(parents=True, exist_ok=True)
OUT_FILE.write_text(json.dumps({
    'version': '2026-02-27',
    'recommendations': recommendations,
    'conceptMeta': concept_meta,
    'missing': missing,
}, indent=2), encoding='utf-8')

TODO_FILE.parent.mkdir(parents=True, exist_ok=True)
if missing:
    TODO_FILE.write_text('# TODO\n\n' + '\n'.join([f'- {m}' for m in sorted(set(missing))]) + '\n', encoding='utf-8')
else:
    TODO_FILE.write_text('# TODO\n\n- No missing mapping rows or files detected.\n', encoding='utf-8')

print(f'Wrote {OUT_FILE}')
print(f'Missing items: {len(set(missing))}')
